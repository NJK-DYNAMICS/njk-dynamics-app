"""
============================================================
 NJK DYNAMICS — NEXUS
 Arquivo: api/nexus/main.py  (v1.3.0)
 Adiciona endpoints do Finalyze:
   POST /api/finalyze/ocr           -> Extrai dados do PDF via Claude Vision
   POST /api/finalyze/contratos     -> Salva contrato no Supabase
   GET  /api/finalyze/contratos     -> Lista contratos
   GET  /api/finalyze/contratos/{id}-> Busca contrato
   POST /api/finalyze/excel/{id}    -> Gera planilha Excel
   POST /api/finalyze/pdf/{id}      -> Gera relatório PDF
============================================================
"""

import os
import io
import base64
import asyncio
import httpx
from datetime import datetime, date
from typing import Optional, List
from math import pow

from fastapi import FastAPI, HTTPException, Query, UploadFile, File
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field, field_validator
from dotenv import load_dotenv
from supabase import create_client, Client

import pdfplumber
import anthropic
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

load_dotenv()

SUPABASE_URL: str = os.environ.get("SUPABASE_URL", "")
SUPABASE_KEY: str = os.environ.get("SUPABASE_KEY", "")
ANTHROPIC_KEY: str = os.environ.get("ANTHROPIC_API_KEY", "")
AGROATA_API_URL:  str = os.environ.get("AGROATA_API_URL", "")
FINALYZE_API_URL: str = os.environ.get("FINALYZE_API_URL", "")

if not SUPABASE_URL or not SUPABASE_KEY:
    raise EnvironmentError("Credenciais Supabase não configuradas.")

supabase: Client = create_client(SUPABASE_URL, SUPABASE_KEY)

app = FastAPI(title="NJK Nexus API", version="1.3.0", docs_url="/docs", redoc_url="/redoc")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ══════════════════════════════════════════════
#  SCHEMAS
# ══════════════════════════════════════════════

STATUS_VALIDOS = {"ativo", "inativo", "suspenso"}

class ClienteCreate(BaseModel):
    nome: str = Field(..., min_length=2, max_length=200)
    status: Optional[str] = Field("ativo")

    @field_validator("nome")
    @classmethod
    def nome_nao_vazio(cls, v):
        if not v.strip(): raise ValueError("Nome vazio.")
        return v.strip()

    @field_validator("status")
    @classmethod
    def status_valido(cls, v):
        if v not in STATUS_VALIDOS: raise ValueError(f"Status inválido.")
        return v

class ClienteStatusUpdate(BaseModel):
    status: str

    @field_validator("status")
    @classmethod
    def status_valido(cls, v):
        if v not in STATUS_VALIDOS: raise ValueError(f"Status inválido.")
        return v

class ContratoCreate(BaseModel):
    cliente_id:      int
    cliente_nome:    str
    numero:          Optional[str] = None
    agente:          Optional[str] = None
    linha_credito:   Optional[str] = None
    data_emissao:    Optional[str] = None
    sistema_amort:   str = "SAC"
    taxa_juros_aa:   float
    valor_total:     float
    n_parcelas:      int
    data_primeira:   str
    correcao_ativa:  bool = False
    indice_correcao: Optional[str] = None
    formula_custom:  Optional[str] = None

# ══════════════════════════════════════════════
#  HELPERS — CÁLCULO DE AMORTIZAÇÃO
# ══════════════════════════════════════════════

def calcular_tabela(valor: float, taxa_aa: float, n: int, primeira: date,
                    sistema: str, correcao: bool, indice: str) -> list:
    taxa_am = taxa_aa / 100 / 12
    hoje = date.today()
    tabela = []
    saldo = valor

    for i in range(1, n + 1):
        # Data de vencimento
        mes = primeira.month + (i - 1)
        ano = primeira.year + (mes - 1) // 12
        mes = ((mes - 1) % 12) + 1
        try:
            venc = date(ano, mes, primeira.day)
        except ValueError:
            import calendar
            ultimo = calendar.monthrange(ano, mes)[1]
            venc = date(ano, mes, ultimo)

        juros = saldo * taxa_am

        if sistema == "SAC":
            principal = valor / n
            prestacao = principal + juros
        elif sistema == "PRICE":
            if taxa_am > 0:
                prestacao = valor * (taxa_am * (1 + taxa_am)**n) / ((1 + taxa_am)**n - 1)
            else:
                prestacao = valor / n
            principal = prestacao - juros
        else:  # CUSTOM — trata como SAC por padrão
            principal = valor / n
            prestacao = principal + juros

        saldo = max(0, saldo - principal)

        # Status
        if venc < hoje:
            status = "liquidada"
        else:
            status = "aberta"

        # Correção para parcelas em aberto vencidas
        prestacao_corrigida = prestacao
        if correcao and venc < hoje:
            meses_atraso = max(0, (hoje.year - venc.year) * 12 + (hoje.month - venc.month))
            if meses_atraso > 0:
                prestacao_corrigida = prestacao * ((1 + taxa_am) ** meses_atraso)
                status = "vencida"

        tabela.append({
            "n": i,
            "vencimento": venc.isoformat(),
            "principal": round(principal, 2),
            "juros": round(juros, 2),
            "prestacao": round(prestacao_corrigida, 2),
            "saldo_devedor": round(saldo, 2),
            "status": status,
        })

    return tabela

def calcular_kpis(tabela: list, valor: float, taxa_aa: float) -> dict:
    abertas = [p for p in tabela if p["status"] != "liquidada"]
    saldo_atual = abertas[0]["saldo_devedor"] + abertas[0]["principal"] if abertas else 0
    total_juros = sum(p["juros"] for p in tabela)
    cet = taxa_aa
    return {
        "saldo_devedor_atual": round(saldo_atual, 2),
        "total_juros": round(total_juros, 2),
        "cet_aa": round(cet, 4),
        "parcelas_restantes": len(abertas),
        "percentual_juros": round((total_juros / valor) * 100, 2) if valor else 0,
    }

# ══════════════════════════════════════════════
#  HELPERS — SUPABASE
# ══════════════════════════════════════════════

async def count_clientes(status: str) -> int:
    try:
        result = supabase.table("dim_clientes").select("id").eq("status", status).execute()
        return len(result.data)
    except Exception:
        return -1

async def fetch_modulo_stats(url: str, path: str):
    if not url: return None
    try:
        async with httpx.AsyncClient(timeout=8.0) as client:
            r = await client.get(f"{url.rstrip('/')}{path}")
            if r.status_code == 200: return r.json()
    except Exception:
        pass
    return None

# ══════════════════════════════════════════════
#  HELPERS — EXCEL
# ══════════════════════════════════════════════

def gerar_excel(contrato: dict, tabela: list, kpis: dict) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Amortização"

    # Cores
    COR_HEADER   = "1A1A2E"
    COR_SUBHEAD  = "16213E"
    COR_ACCENT   = "F59E0B"
    COR_ROW_ALT  = "F8F9FA"
    COR_LIQ      = "D1FAE5"
    COR_ABERTA   = "FEF3C7"
    COR_VENCIDA  = "FEE2E2"
    BRANCO       = "FFFFFF"

    fonte_titulo  = Font(name="Calibri", bold=True, size=16, color=BRANCO)
    fonte_header  = Font(name="Calibri", bold=True, size=10, color=BRANCO)
    fonte_label   = Font(name="Calibri", bold=True, size=10, color="374151")
    fonte_valor   = Font(name="Calibri", size=10, color="111827")
    fonte_mono    = Font(name="Courier New", size=9)

    def fill(hex_color):
        return PatternFill("solid", fgColor=hex_color)

    def border_thin():
        s = Side(style="thin", color="E5E7EB")
        return Border(left=s, right=s, top=s, bottom=s)

    # ── TÍTULO
    ws.merge_cells("A1:G1")
    ws["A1"] = "NJK DYNAMICS — FINALYZE"
    ws["A1"].font = Font(name="Calibri", bold=True, size=18, color=COR_ACCENT)
    ws["A1"].fill = fill(COR_HEADER)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    ws.merge_cells("A2:G2")
    ws["A2"] = "RELATÓRIO DE AMORTIZAÇÃO DE FINANCIAMENTO"
    ws["A2"].font = Font(name="Calibri", bold=True, size=11, color=BRANCO)
    ws["A2"].fill = fill(COR_SUBHEAD)
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 22

    # ── CABEÇALHO DO CONTRATO
    row = 4
    dados_cab = [
        ("Tomador / Cliente:", contrato.get("cliente_nome","—")),
        ("Agente Financeiro:",  contrato.get("agente","—")),
        ("Linha de Crédito:",   contrato.get("linha_credito","—")),
        ("Nº do Contrato:",     contrato.get("numero","—")),
        ("Data de Emissão:",    contrato.get("data_emissao","—")),
        ("Sistema de Amort.:",  contrato.get("sistema_amort","—")),
        ("Taxa de Juros:",      f"{contrato.get('taxa_juros_aa',0):.2f}% a.a."),
        ("Valor Financiado:",   f"R$ {contrato.get('valor_total',0):,.2f}"),
        ("Nº de Parcelas:",     str(contrato.get("n_parcelas",0))),
        ("Data 1ª Parcela:",    contrato.get("data_primeira","—")),
        ("Data do Relatório:",  datetime.now().strftime("%d/%m/%Y")),
    ]

    for label, valor in dados_cab:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = fonte_label
        ws[f"A{row}"].fill = fill("F3F4F6")
        ws.merge_cells(f"B{row}:D{row}")
        ws[f"B{row}"] = valor
        ws[f"B{row}"].font = fonte_valor
        row += 1

    # ── KPIs
    row += 1
    ws.merge_cells(f"A{row}:G{row}")
    ws[f"A{row}"] = "INDICADORES FINANCEIROS"
    ws[f"A{row}"].font = fonte_header
    ws[f"A{row}"].fill = fill(COR_HEADER)
    ws[f"A{row}"].alignment = Alignment(horizontal="center")
    row += 1

    kpi_dados = [
        ("Saldo Devedor Atual", f"R$ {kpis['saldo_devedor_atual']:,.2f}"),
        ("Total de Juros",      f"R$ {kpis['total_juros']:,.2f}"),
        ("CET (a.a.)",          f"{kpis['cet_aa']:.2f}%"),
        ("Parcelas Restantes",  str(kpis['parcelas_restantes'])),
        ("% Juros / Financi.",  f"{kpis['percentual_juros']:.1f}%"),
    ]
    for label, val in kpi_dados:
        ws[f"A{row}"] = label; ws[f"A{row}"].font = fonte_label; ws[f"A{row}"].fill = fill("FEF3C7")
        ws[f"B{row}"] = val;   ws[f"B{row}"].font = Font(name="Calibri", bold=True, size=11, color="92400E")
        row += 1

    # ── TABELA DE AMORTIZAÇÃO
    row += 1
    ws.merge_cells(f"A{row}:G{row}")
    ws[f"A{row}"] = "TABELA DE AMORTIZAÇÃO"
    ws[f"A{row}"].font = fonte_header
    ws[f"A{row}"].fill = fill(COR_HEADER)
    ws[f"A{row}"].alignment = Alignment(horizontal="center")
    row += 1

    headers = ["#", "Vencimento", "Principal (R$)", "Juros (R$)", "Prestação (R$)", "Saldo Devedor (R$)", "Status"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = fonte_header
        cell.fill = fill(COR_SUBHEAD)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border_thin()
    row += 1

    for p in tabela:
        cor_row = COR_LIQ if p["status"] == "liquidada" else COR_VENCIDA if p["status"] == "vencida" else COR_ABERTA
        valores = [p["n"], p["vencimento"], p["principal"], p["juros"], p["prestacao"], p["saldo_devedor"], p["status"].upper()]
        for col, val in enumerate(valores, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.fill = fill(cor_row)
            cell.font = fonte_mono
            cell.border = border_thin()
            if col in [3,4,5,6]:
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal="right")
            elif col == 1:
                cell.alignment = Alignment(horizontal="center")
        row += 1

    # Totais
    total_principal = sum(p["principal"] for p in tabela)
    total_juros_    = sum(p["juros"]     for p in tabela)
    total_prestacao = sum(p["prestacao"] for p in tabela)
    totais = ["TOTAL", "", total_principal, total_juros_, total_prestacao, "", ""]
    for col, val in enumerate(totais, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.font = Font(name="Calibri", bold=True, size=10)
        cell.fill = fill("E5E7EB")
        cell.border = border_thin()
        if col in [3,4,5]: cell.number_format = '#,##0.00'; cell.alignment = Alignment(horizontal="right")

    # Larguras
    larguras = [6, 14, 16, 16, 16, 18, 14]
    for i, w in enumerate(larguras, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()

# ══════════════════════════════════════════════
#  HELPERS — PDF
# ══════════════════════════════════════════════

def gerar_pdf(contrato: dict, tabela: list, kpis: dict) -> bytes:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
        rightMargin=1.5*cm, leftMargin=1.5*cm, topMargin=1.5*cm, bottomMargin=1.5*cm)

    AMBER  = colors.HexColor("#F59E0B")
    DARK   = colors.HexColor("#1A1A2E")
    GRAY   = colors.HexColor("#374151")
    LGRAY  = colors.HexColor("#F3F4F6")
    GREEN  = colors.HexColor("#D1FAE5")
    YELLOW = colors.HexColor("#FEF3C7")
    RED    = colors.HexColor("#FEE2E2")
    WHITE  = colors.white

    styles = getSampleStyleSheet()
    s_title = ParagraphStyle("title", fontName="Helvetica-Bold", fontSize=18, textColor=AMBER, spaceAfter=4, alignment=TA_CENTER)
    s_sub   = ParagraphStyle("sub",   fontName="Helvetica",      fontSize=10, textColor=GRAY,  spaceAfter=2, alignment=TA_CENTER)
    s_label = ParagraphStyle("label", fontName="Helvetica-Bold", fontSize=9,  textColor=GRAY)
    s_small = ParagraphStyle("small", fontName="Helvetica",      fontSize=8,  textColor=GRAY)
    s_sec   = ParagraphStyle("sec",   fontName="Helvetica-Bold", fontSize=11, textColor=WHITE, backColor=DARK, spaceAfter=6, spaceBefore=12, leftIndent=6)

    story = []

    # CABEÇALHO
    story.append(Paragraph("NJK DYNAMICS — FINALYZE", s_title))
    story.append(Paragraph("Relatório de Amortização de Financiamento", s_sub))
    story.append(Spacer(1, 0.3*cm))

    # DADOS DO CONTRATO
    story.append(Paragraph("DADOS DO CONTRATO", s_sec))
    cab_data = [
        ["Tomador / Cliente", contrato.get("cliente_nome","—"),   "Nº do Contrato", contrato.get("numero","—")],
        ["Agente Financeiro", contrato.get("agente","—"),         "Data de Emissão", contrato.get("data_emissao","—")],
        ["Linha de Crédito",  contrato.get("linha_credito","—"),  "Sistema de Amort.", contrato.get("sistema_amort","—")],
        ["Valor Financiado",  f"R$ {contrato.get('valor_total',0):,.2f}", "Taxa de Juros", f"{contrato.get('taxa_juros_aa',0):.2f}% a.a."],
        ["Nº de Parcelas",    str(contrato.get("n_parcelas",0)),  "1ª Parcela", contrato.get("data_primeira","—")],
        ["Data do Relatório", datetime.now().strftime("%d/%m/%Y %H:%M"), "", ""],
    ]

    t_cab = Table(cab_data, colWidths=[3.5*cm, 6*cm, 3.5*cm, 5*cm])
    t_cab.setStyle(TableStyle([
        ('FONTNAME',  (0,0),(-1,-1), 'Helvetica'),
        ('FONTSIZE',  (0,0),(-1,-1), 8),
        ('FONTNAME',  (0,0),(0,-1),  'Helvetica-Bold'),
        ('FONTNAME',  (2,0),(2,-1),  'Helvetica-Bold'),
        ('BACKGROUND',(0,0),(0,-1),  LGRAY),
        ('BACKGROUND',(2,0),(2,-1),  LGRAY),
        ('GRID',      (0,0),(-1,-1), 0.5, colors.HexColor("#E5E7EB")),
        ('PADDING',   (0,0),(-1,-1), 5),
        ('VALIGN',    (0,0),(-1,-1), 'MIDDLE'),
    ]))
    story.append(t_cab)
    story.append(Spacer(1, 0.3*cm))

    # KPIs
    story.append(Paragraph("INDICADORES FINANCEIROS", s_sec))
    kpi_data = [
        ["Saldo Devedor Atual", f"R$ {kpis['saldo_devedor_atual']:,.2f}",
         "Total de Juros", f"R$ {kpis['total_juros']:,.2f}",
         "CET (a.a.)", f"{kpis['cet_aa']:.2f}%"],
    ]
    t_kpi = Table(kpi_data, colWidths=[3.5*cm, 4.5*cm, 3*cm, 4.5*cm, 2.5*cm, 2.5*cm])
    t_kpi.setStyle(TableStyle([
        ('FONTNAME',  (0,0),(-1,-1), 'Helvetica'),
        ('FONTNAME',  (0,0),(0,-1),  'Helvetica-Bold'),
        ('FONTNAME',  (2,0),(2,-1),  'Helvetica-Bold'),
        ('FONTNAME',  (4,0),(4,-1),  'Helvetica-Bold'),
        ('FONTSIZE',  (0,0),(-1,-1), 9),
        ('FONTSIZE',  (1,0),(1,0),   12),
        ('FONTSIZE',  (3,0),(3,0),   12),
        ('FONTSIZE',  (5,0),(5,0),   12),
        ('BACKGROUND',(0,0),(0,-1),  YELLOW),
        ('BACKGROUND',(2,0),(2,-1),  YELLOW),
        ('BACKGROUND',(4,0),(4,-1),  YELLOW),
        ('TEXTCOLOR', (1,0),(1,0),   colors.HexColor("#92400E")),
        ('TEXTCOLOR', (3,0),(3,0),   colors.HexColor("#92400E")),
        ('TEXTCOLOR', (5,0),(5,0),   colors.HexColor("#92400E")),
        ('GRID',      (0,0),(-1,-1), 0.5, colors.HexColor("#E5E7EB")),
        ('PADDING',   (0,0),(-1,-1), 8),
        ('ALIGN',     (1,0),(1,0),   'RIGHT'),
        ('ALIGN',     (3,0),(3,0),   'RIGHT'),
        ('ALIGN',     (5,0),(5,0),   'RIGHT'),
    ]))
    story.append(t_kpi)
    story.append(Spacer(1, 0.3*cm))

    # TABELA DE AMORTIZAÇÃO
    story.append(Paragraph("TABELA DE AMORTIZAÇÃO", s_sec))
    headers = ["#", "Vencimento", "Principal", "Juros", "Prestação", "Saldo Devedor", "Status"]
    t_data = [headers]
    for p in tabela:
        t_data.append([
            str(p["n"]),
            p["vencimento"],
            f"R$ {p['principal']:,.2f}",
            f"R$ {p['juros']:,.2f}",
            f"R$ {p['prestacao']:,.2f}",
            f"R$ {p['saldo_devedor']:,.2f}",
            p["status"].upper(),
        ])
    # Totais
    t_data.append([
        "TOTAL", "",
        f"R$ {sum(p['principal'] for p in tabela):,.2f}",
        f"R$ {sum(p['juros']     for p in tabela):,.2f}",
        f"R$ {sum(p['prestacao'] for p in tabela):,.2f}",
        "", "",
    ])

    col_w = [1*cm, 2.5*cm, 3*cm, 3*cm, 3*cm, 3.5*cm, 2*cm]
    t_amort = Table(t_data, colWidths=col_w, repeatRows=1)

    style_amort = [
        ('FONTNAME',   (0,0),(-1,0),   'Helvetica-Bold'),
        ('FONTSIZE',   (0,0),(-1,-1),  8),
        ('BACKGROUND', (0,0),(-1,0),   DARK),
        ('TEXTCOLOR',  (0,0),(-1,0),   WHITE),
        ('ALIGN',      (0,0),(-1,-1),  'RIGHT'),
        ('ALIGN',      (0,0),(0,-1),   'CENTER'),
        ('ALIGN',      (1,0),(1,-1),   'CENTER'),
        ('ALIGN',      (6,0),(6,-1),   'CENTER'),
        ('GRID',       (0,0),(-1,-1),  0.4, colors.HexColor("#E5E7EB")),
        ('PADDING',    (0,0),(-1,-1),  4),
        # Totais
        ('FONTNAME',   (0,-1),(-1,-1), 'Helvetica-Bold'),
        ('BACKGROUND', (0,-1),(-1,-1), colors.HexColor("#E5E7EB")),
    ]

    # Cores por status
    for i, p in enumerate(tabela, 1):
        if p["status"] == "liquidada":
            style_amort.append(('BACKGROUND', (0,i),(-1,i), GREEN))
        elif p["status"] == "vencida":
            style_amort.append(('BACKGROUND', (0,i),(-1,i), RED))
        else:
            style_amort.append(('BACKGROUND', (0,i),(-1,i), YELLOW))

    t_amort.setStyle(TableStyle(style_amort))
    story.append(t_amort)

    # RODAPÉ
    story.append(Spacer(1, 0.5*cm))
    story.append(Paragraph(
        f"Relatório gerado em {datetime.now().strftime('%d/%m/%Y às %H:%M')} · NJK Dynamics — Finalyze · njkdynamics.vercel.app",
        ParagraphStyle("footer", fontName="Helvetica", fontSize=7, textColor=colors.HexColor("#9CA3AF"), alignment=TA_CENTER)
    ))

    doc.build(story)
    buf.seek(0)
    return buf.read()

# ══════════════════════════════════════════════
#  ENDPOINTS — SISTEMA
# ══════════════════════════════════════════════

@app.get("/health", tags=["Sistema"])
async def health_check():
    return {"status":"online","servico":"NJK Nexus API","versao":"1.3.0","timestamp":datetime.utcnow().isoformat()+"Z"}

# ══════════════════════════════════════════════
#  ENDPOINTS — CLIENTES (CRUD)
# ══════════════════════════════════════════════

@app.post("/api/clientes", status_code=201, tags=["Clientes"])
async def criar_cliente(payload: ClienteCreate):
    try:
        r = supabase.table("dim_clientes").insert({"nome":payload.nome,"status":payload.status}).execute()
        if r.data: return {"sucesso":True,"mensagem":"Cliente cadastrado.","cliente":r.data[0]}
        raise HTTPException(500,"Erro ao inserir.")
    except HTTPException: raise
    except Exception as e: raise HTTPException(500,str(e))

@app.get("/api/clientes", tags=["Clientes"])
async def listar_clientes(
    status: Optional[str] = Query(None),
    limite: int = Query(100, ge=1, le=9999),
    pagina: int = Query(1, ge=1),
):
    try:
        q = supabase.table("dim_clientes").select("*")
        fs = status if status else "ativo"
        if fs in STATUS_VALIDOS: q = q.eq("status", fs)
        offset = (pagina-1)*limite
        r = q.order("id",desc=False).range(offset, offset+limite-1).execute()
        return {"sucesso":True,"total":len(r.data),"pagina":pagina,"limite":limite,"filtro_status":fs,"clientes":r.data}
    except Exception as e: raise HTTPException(500,str(e))

@app.get("/api/clientes/{cliente_id}", tags=["Clientes"])
async def buscar_cliente(cliente_id: int):
    try:
        r = supabase.table("dim_clientes").select("*").eq("id",cliente_id).single().execute()
        if r.data: return {"sucesso":True,"cliente":r.data}
        raise HTTPException(404,"Cliente não encontrado.")
    except HTTPException: raise
    except Exception as e: raise HTTPException(500,str(e))

@app.patch("/api/clientes/{cliente_id}/status", tags=["Clientes"])
async def atualizar_status_cliente(cliente_id: int, payload: ClienteStatusUpdate):
    try:
        r = supabase.table("dim_clientes").update({"status":payload.status}).eq("id",cliente_id).execute()
        if r.data: return {"sucesso":True,"mensagem":f"Status -> '{payload.status}'.","cliente":r.data[0]}
        raise HTTPException(404,"Cliente não encontrado.")
    except HTTPException: raise
    except Exception as e: raise HTTPException(500,str(e))

@app.delete("/api/clientes/{cliente_id}", tags=["Clientes"])
async def excluir_cliente(cliente_id: int):
    try:
        r = supabase.table("dim_clientes").delete().eq("id",cliente_id).execute()
        if r.data: return {"sucesso":True,"mensagem":f"Cliente #{cliente_id} removido."}
        raise HTTPException(404,"Cliente não encontrado.")
    except HTTPException: raise
    except Exception as e: raise HTTPException(500,str(e))

# ══════════════════════════════════════════════
#  ENDPOINTS — FINALYZE
# ══════════════════════════════════════════════

@app.post("/api/finalyze/ocr", tags=["Finalyze"])
async def extrair_pdf(file: UploadFile = File(...)):
    """Extrai dados de contrato de um PDF via Claude Vision."""
    if not ANTHROPIC_KEY:
        raise HTTPException(501, "ANTHROPIC_API_KEY não configurada.")
    if not file.filename.lower().endswith(".pdf"):
        raise HTTPException(400, "Apenas arquivos PDF são aceitos.")
    try:
        content = await file.read()

        # Tenta extrair texto com pdfplumber primeiro
        texto = ""
        with pdfplumber.open(io.BytesIO(content)) as pdf:
            for page in pdf.pages:
                texto += (page.extract_text() or "") + "\n"

        # Envia para Claude extrair campos estruturados
        client_ai = anthropic.Anthropic(api_key=ANTHROPIC_KEY)
        prompt = f"""Analise o texto abaixo de um contrato/cédula de crédito bancário e extraia os campos em JSON.
Retorne APENAS o JSON, sem explicações, sem markdown.

Campos esperados:
- numero: número do contrato
- agente: nome do banco/agente financeiro
- linha_credito: linha de crédito (ex: PRONAF, MODERFROTA, etc)
- data_emissao: data de emissão (formato YYYY-MM-DD)
- sistema_amort: sistema de amortização (SAC, PRICE ou outro)
- taxa_juros_aa: taxa de juros anual como número (ex: 7.5)
- valor_total: valor financiado como número (ex: 500000.00)
- n_parcelas: número de parcelas como inteiro
- data_primeira: data da primeira parcela (formato YYYY-MM-DD)

Se algum campo não for encontrado, use null.

TEXTO DO CONTRATO:
{texto[:6000]}"""

        msg = client_ai.messages.create(
            model="claude-sonnet-4-20250514",
            max_tokens=1000,
            messages=[{"role":"user","content":prompt}]
        )

        import json
        raw = msg.content[0].text.strip()
        dados = json.loads(raw)
        return {"sucesso": True, "dados": dados}

    except Exception as e:
        raise HTTPException(500, f"Erro ao processar PDF: {str(e)}")


@app.post("/api/finalyze/contratos", status_code=201, tags=["Finalyze"])
async def salvar_contrato(payload: ContratoCreate):
    """Salva contrato no Supabase e retorna a tabela de amortização calculada."""
    try:
        primeira = date.fromisoformat(payload.data_primeira)
        tabela = calcular_tabela(
            payload.valor_total, payload.taxa_juros_aa, payload.n_parcelas,
            primeira, payload.sistema_amort, payload.correcao_ativa, payload.indice_correcao or ""
        )
        kpis = calcular_kpis(tabela, payload.valor_total, payload.taxa_juros_aa)

        insert_data = payload.model_dump()
        r = supabase.table("dim_contratos").insert(insert_data).execute()
        if not r.data:
            raise HTTPException(500, "Erro ao inserir contrato.")

        contrato_id = r.data[0]["id"]
        return {
            "sucesso": True,
            "contrato_id": contrato_id,
            "contrato": r.data[0],
            "tabela": tabela,
            "kpis": kpis,
        }
    except HTTPException: raise
    except Exception as e: raise HTTPException(500, str(e))


@app.get("/api/finalyze/contratos", tags=["Finalyze"])
async def listar_contratos(
    cliente_id: Optional[int] = Query(None),
    limite: int = Query(50, ge=1, le=200),
):
    try:
        q = supabase.table("dim_contratos").select("*").order("criado_em", desc=True).limit(limite)
        if cliente_id: q = q.eq("cliente_id", cliente_id)
        r = q.execute()
        return {"sucesso": True, "total": len(r.data), "contratos": r.data}
    except Exception as e: raise HTTPException(500, str(e))


@app.get("/api/finalyze/contratos/{contrato_id}", tags=["Finalyze"])
async def buscar_contrato(contrato_id: int):
    try:
        r = supabase.table("dim_contratos").select("*").eq("id", contrato_id).single().execute()
        if not r.data: raise HTTPException(404, "Contrato não encontrado.")
        contrato = r.data
        primeira = date.fromisoformat(contrato["data_primeira"])
        tabela = calcular_tabela(
            contrato["valor_total"], contrato["taxa_juros_aa"], contrato["n_parcelas"],
            primeira, contrato["sistema_amort"], contrato["correcao_ativa"], contrato.get("indice_correcao","")
        )
        kpis = calcular_kpis(tabela, contrato["valor_total"], contrato["taxa_juros_aa"])
        return {"sucesso": True, "contrato": contrato, "tabela": tabela, "kpis": kpis}
    except HTTPException: raise
    except Exception as e: raise HTTPException(500, str(e))


@app.post("/api/finalyze/excel/{contrato_id}", tags=["Finalyze"])
async def exportar_excel(contrato_id: int):
    """Gera planilha Excel do contrato e retorna para download."""
    try:
        r = supabase.table("dim_contratos").select("*").eq("id", contrato_id).single().execute()
        if not r.data: raise HTTPException(404, "Contrato não encontrado.")
        contrato = r.data
        primeira = date.fromisoformat(contrato["data_primeira"])
        tabela = calcular_tabela(
            contrato["valor_total"], contrato["taxa_juros_aa"], contrato["n_parcelas"],
            primeira, contrato["sistema_amort"], contrato["correcao_ativa"], contrato.get("indice_correcao","")
        )
        kpis = calcular_kpis(tabela, contrato["valor_total"], contrato["taxa_juros_aa"])
        xlsx_bytes = gerar_excel(contrato, tabela, kpis)
        filename = f"finalyze_contrato_{contrato_id}_{datetime.now().strftime('%Y%m%d')}.xlsx"
        return StreamingResponse(
            io.BytesIO(xlsx_bytes),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except HTTPException: raise
    except Exception as e: raise HTTPException(500, str(e))


@app.post("/api/finalyze/pdf/{contrato_id}", tags=["Finalyze"])
async def exportar_pdf(contrato_id: int):
    """Gera relatório PDF do contrato e retorna para download."""
    try:
        r = supabase.table("dim_contratos").select("*").eq("id", contrato_id).single().execute()
        if not r.data: raise HTTPException(404, "Contrato não encontrado.")
        contrato = r.data
        primeira = date.fromisoformat(contrato["data_primeira"])
        tabela = calcular_tabela(
            contrato["valor_total"], contrato["taxa_juros_aa"], contrato["n_parcelas"],
            primeira, contrato["sistema_amort"], contrato["correcao_ativa"], contrato.get("indice_correcao","")
        )
        kpis = calcular_kpis(tabela, contrato["valor_total"], contrato["taxa_juros_aa"])
        pdf_bytes = gerar_pdf(contrato, tabela, kpis)
        filename = f"finalyze_relatorio_{contrato_id}_{datetime.now().strftime('%Y%m%d')}.pdf"
        return StreamingResponse(
            io.BytesIO(pdf_bytes),
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except HTTPException: raise
    except Exception as e: raise HTTPException(500, str(e))

# ══════════════════════════════════════════════
#  ENDPOINTS — STATS (NEXUS Portal)
# ══════════════════════════════════════════════

@app.get("/api/stats/dataware", tags=["Stats"])
async def stats_dataware():
    ativos, inativos, suspensos = await asyncio.gather(
        count_clientes("ativo"), count_clientes("inativo"), count_clientes("suspenso"))
    erro = any(v == -1 for v in [ativos, inativos, suspensos])
    total = max(ativos,0)+max(inativos,0)+max(suspensos,0)
    return {"sucesso":not erro,"total":total,"ativos":ativos,"inativos":inativos,"suspensos":suspensos,"outros":max(inativos,0)+max(suspensos,0),"timestamp":datetime.utcnow().isoformat()+"Z"}

@app.get("/api/stats/agroata", tags=["Stats"])
async def stats_agroata():
    if not AGROATA_API_URL: return {"sucesso":False,"disponivel":False,"mensagem":"AgroAta não configurado.","total":None,"produtores":None,"mes":None}
    dados = await fetch_modulo_stats(AGROATA_API_URL, "/api/atas/stats")
    if dados: return {"sucesso":True,"disponivel":True,"total":dados.get("total"),"produtores":dados.get("produtores"),"mes":dados.get("mes"),"timestamp":datetime.utcnow().isoformat()+"Z"}
    return {"sucesso":False,"disponivel":True,"mensagem":"AgroAta indisponível.","total":None,"produtores":None,"mes":None}

@app.get("/api/stats/finalyze", tags=["Stats"])
async def stats_finalyze():
    try:
        r = supabase.table("dim_contratos").select("id").execute()
        total = len(r.data)
        return {"sucesso":True,"disponivel":True,"contratos":total,"pdfs":total,"relatorios":total,"timestamp":datetime.utcnow().isoformat()+"Z"}
    except:
        return {"sucesso":False,"disponivel":False,"contratos":None,"pdfs":None,"relatorios":None}

@app.get("/api/stats", tags=["Stats"])
async def stats_completas():
    dataware, agroata, finalyze = await asyncio.gather(stats_dataware(), stats_agroata(), stats_finalyze())
    return {"sucesso":True,"timestamp":datetime.utcnow().isoformat()+"Z","modulos":{"dataware":dataware,"agroata":agroata,"finalyze":finalyze}}
